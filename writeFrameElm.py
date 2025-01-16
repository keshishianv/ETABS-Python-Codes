from openpyxl import Workbook, load_workbook

def update_excel_with_FrameElm(file_name, data):
    """
    Creates or updates an Excel file with a tab 'Get Frame Members' and populates it with data.

    Args:
        file_name (str): Name of the Excel file.
        data (list): Data to populate the spreadsheet. Each element in the list can be:
                     - A float, str, or other single value: It will be treated as a single-column row.
                     - A list or tuple: It will be treated as a multi-column row.
    """
    # Column titles
    headers = [
        "Frame Name", "Unknown 1", "Unknown 2", "Unknown 3", "Unknown 4", "Unknown 5"
    ]

    # Try to load the existing workbook or create a new one
    try:
        workbook = load_workbook(file_name)
    except FileNotFoundError:
        workbook = Workbook()

    # Check if the sheet exists, otherwise create it
    if "Get Frame ELM" in workbook.sheetnames:
        sheet = workbook["Get Frame ELM"]
        # Clear the existing sheet
        workbook.remove(sheet)
        sheet = workbook.create_sheet("Get Frame ELM")
    else:
        sheet = workbook.create_sheet("Get Frame ELM")

    # Add headers to the first row
    sheet.append(headers)

    row = 2
    for kk in data:
        colInt = 1
        for i in kk:
            for k, value in enumerate(i, start=2):
                sheet.cell(row=row + k, column=colInt, value=value)
            colInt += 1
        row = row + len(kk)

    # Save the workbook
    workbook.save(file_name)
    print(f"Data successfully written to {file_name} in 'Get Frame ELM' sheet.")
