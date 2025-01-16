from openpyxl import Workbook, load_workbook

def update_excel_with_data(file_name, data):
    """
    Creates or updates an Excel file with a tab 'Get Frame Members' and populates it with data.

    Args:
        file_name (str): Name of the Excel file.
        data (list): Data to populate the spreadsheet. Each element in the list can be:
                     - A float, str, or other single value: It will be treated as a single-column row.
                     - A list or tuple: It will be treated as a multi-column row.
    """
    # Column titles
    headers = [
        "NumberNames", "MyName", "PropName", "StoryName", "PointName1", "PointName2",
        "Point1X", "Point1Y", "Point1Z", "Point2X", "Point2Y", "Point2Z", "Angle",
        "Offset1X", "Offset2X", "Offset1Y", "Offset2Y", "Offset1Z", "Offset2Z",
        "CardinalPoint", "Global"
    ]

    # Try to load the existing workbook or create a new one
    try:
        workbook = load_workbook(file_name)
    except FileNotFoundError:
        workbook = Workbook()

    # Check if the sheet exists, otherwise create it
    if "Get Frame Members" in workbook.sheetnames:
        sheet = workbook["Get Frame Members"]
        # Clear the existing sheet
        workbook.remove(sheet)
        sheet = workbook.create_sheet("Get Frame Members")
    else:
        sheet = workbook.create_sheet("Get Frame Members")

    # Add headers to the first row
    sheet.append(headers)

    # Add data rows
    colInt = 1
    for i in data:
        for k, value in enumerate(i,start=2):
            sheet.cell(row=k, column=colInt, value=value)
        colInt+=1


    # Save the workbook
    workbook.save(file_name)
    print(f"Data successfully written to {file_name} in 'Get Frame Members' sheet.")
