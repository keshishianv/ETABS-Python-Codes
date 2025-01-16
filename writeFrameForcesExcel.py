from openpyxl import Workbook, load_workbook

def update_excel_with_frame_forces_data(file_name, data):
    """
    Creates or updates an Excel file with a tab 'Get Frame Members' and populates it with data.

    Args:
        file_name (str): Name of the Excel file.
        data (list): Data to populate the spreadsheet. Each element in the list can be:
                     - A float, str, or other single value: It will be treated as a single-column row.
                     - A list or tuple: It will be treated as a multi-column row.
    """
    # Column titles
    headers = [
        "Cases Names","Combos Names","NumberResults", "Unique Frame #", "i end (ft)", "Elm name", "i end (ft)", "LoadCase","LoadCombos", "StepType", "StepNum", "P",
        "V2(i)", "V3(i)", "T(i) / 12", "M2(i) / 12", "M3"
    ]

    # Try to load the existing workbook or create a new one
    try:
        workbook = load_workbook(file_name)
    except FileNotFoundError:
        workbook = Workbook()

    # Check if the sheet exists, otherwise create it
    if "Get Frame Forces" in workbook.sheetnames:
        sheet2 = workbook["Get Frame Forces"]
        # Clear the existing sheet
        workbook.remove(sheet2)
        sheet2 = workbook.create_sheet("Get Frame Forces")
    else:
        sheet2 = workbook.create_sheet("Get Frame Forces")

    # Add headers to the first row
    sheet2.append(headers)

    # Add data rows
    colInt = 1
    for i in data:
        for k, value in enumerate(i,start=2):
            sheet2.cell(row=k, column=colInt, value=value)
        colInt+=1



    # Save the workbook
    workbook.save(file_name)
    print(f"Data successfully written to {file_name} in 'Get Frame Forces' sheet.")
